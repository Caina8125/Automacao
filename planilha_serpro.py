from pandas import DataFrame
from pandas import ExcelWriter
from pandas import read_excel
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class PlanilhaSerpro(): 
    def __init__(self, path_planilha: str) -> None:
        self._df_planilha: DataFrame = read_excel(path_planilha).loc[:, [
            'Fatura', 'Protocolo Glosa', 'Valor Recursado', 'Controle Inicial', 'Nro. Guia', 'Autorização', 'Matrícula', 'Paciente', 'Procedimento',
            'Descrição','Valor Original', 'Usuário Glosa', 'Valor Glosa', 'Motivo Glosa', 'Valor Cobrado', 'Recurso Glosa'
            ]]
        self._book: Workbook = load_workbook(r"C:\Users\lucas.paz\Documents\Serpro\RECURSO DE GLOSAS SERPRO - DEZ 2023 .xlsx")
        self.sheet = self._book['Recurso de Glosas Serpro']
    
    def create_writer(self, numero_processo: str) -> ExcelWriter:
        writer: ExcelWriter = ExcelWriter(f'output\\Serpro_{numero_processo}.xlsx')
        writer.book = self._book
        writer.sheets = dict((ws.title, ws) for ws in self._book.worksheets)
        return writer
    
    def filter_df_by_number(self, numero_processo) -> DataFrame:
        df_processo: DataFrame = self._df_planilha.loc[(self._df_planilha["Fatura"] == int(numero_processo))]

        if df_processo.empty:
            return self._df_planilha.loc[(self._df_planilha["Fatura"] == numero_processo)]
        
        return df_processo
    
    def get_info_processo(self, df_processo: DataFrame) -> None:
        print(df_processo['Protocolo Glosa'])
        return {
            'protocolo': f'{df_processo["Protocolo Glosa"][0]}'.replace('.0', ''),
            'numero_fatura': f'{df_processo["Fatura"][0]}',
            'valor_total_original': f'{df_processo["Valor Original"].sum()}',
            'valor_liberado': f'{df_processo["Valor Original"].sum() + df_processo["Valor Glosa"].sum()}',
            'valor_glosa_total': f'{df_processo["Valor Glosa"].sum()}',
            'valor_recurso_total': f'{df_processo["Valor Recursado"].sum()}'
        }
    
    def atualiza_template(self, chave, dado) -> None:
        match chave:
            case 'protocolo':
                self.add_info_na_celula('Protocolo:', dado)

            case 'numero_fatura':
                self.add_info_na_celula('Lote:', dado)

            case 'valor_total_original':
                self.add_info_na_celula('Valor Informado:', dado)

            case 'valor_liberado':
                self.add_info_na_celula('Valor Liberado:', dado)

            case 'valor_glosa_total':
                self.add_info_na_celula('Valor Glosa:', dado)

            case 'valor_recurso_total':
                self.add_info_na_celula('Valor Recurso:', dado)
    
    def add_info_na_celula(self, valor_celula: str, valor: str) -> None:
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == valor_celula:
                    new_value = valor_celula + ' ' + valor
                    self.sheet.cell(row=cell.row, column=cell.column).value = new_value
    
    def create_excel(self, writer: ExcelWriter, df_processo: DataFrame) -> None:
        df_processo = df_processo.drop(['Fatura', 'Protocolo Glosa', 'Valor Recursado'], axis='columns')
        df_processo.to_excel(writer, 'Recurso de Glosas Serpro', startrow=5, startcol=0, header=False, index=False)

    def a_ser_nomeada(self) -> None:
        LISTA_DE_PROCESSOS: list[str] = [
            f'{value}'.replace('.0', '') 
            for value in list(set(self._df_planilha['Fatura'].values.tolist()))
            ]

        for numero_processo in LISTA_DE_PROCESSOS:
            df_processo: DataFrame = self.filter_df_by_number(numero_processo)
            print(df_processo.columns)
            info_processo: dict = self.get_info_processo(df_processo)

            for chave, dado in info_processo.items(): 
                self.atualiza_template(chave, dado)

            writer: ExcelWriter = self.create_writer(numero_processo)
            self.create_excel(writer, df_processo)
            writer.close()

teste = PlanilhaSerpro(path_planilha=r"C:\Users\lucas.paz\Documents\Serpro\SERPRO.xlsx")
teste.a_ser_nomeada()