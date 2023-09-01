import pandas as pd
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook
from abc import ABC
import time
from tkinter import filedialog
from selenium.webdriver.chrome.options import Options
from seleniumwire import webdriver
import tkinter

class PageElement(ABC):
    def __init__(self,driver,url=''):
        self.driver = driver
        self.url = url
    def open(self):
        self.driver.get(self.url)


class Login(PageElement):
    multiusuario = (By.XPATH,"/html/body/div[3]/div[3]/div/form/div[1]/label")
    login = (By.XPATH,"/html/body/div[3]/div[3]/div/form/div[2]/div[1]/div/input")
    senha = (By.XPATH,"/html/body/div[3]/div[3]/div/form/div[3]/input")
    cpf = (By.XPATH,"/html/body/div[3]/div[3]/div/form/div[2]/div[2]/div/input")
    logar = (By.XPATH,"/html/body/div[3]/div[3]/div/form/div[4]")

    def exe_login(self, login, senha, cpf):
        self.driver.find_element(*self.multiusuario).click()
        self.driver.find_element(*self.login).send_keys(login)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.cpf).send_keys(cpf)
        self.driver.find_element(*self.logar).click()

        
class caminho(PageElement):
    alerta = (By.XPATH,' /html/body/div[2]/div/center/a')
    guia = (By.XPATH,'//*[@id="objTableDetalhe"]/tbody/tr[3]/td[1]/a')
    envio_xml = (By.XPATH,'//*[@id="main"]/div/div/div[2]/div[2]/article/div[6]/div[4]/div[4]/div[4]/div/div[2]/ul/li[2]/a')
    sem_erros = (By.XPATH,'//*[@id="StaErro"]/option[2]')
    listar = (By.XPATH,'//*[@id="MenuOptionReport"]')

    def exe_caminho(self):
        time.sleep(4)
        try:
            self.driver.find_element(*self.alerta).click()
        except:
            print('Alerta não apareceu')

        self.driver.get("https://www2.geap.com.br/PRESTADOR/portal-tiss.asp#")
        time.sleep(2)
        self.driver.find_element(*self.envio_xml).click()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[1])
        time.sleep(1)
        self.driver.find_element(*self.sem_erros).click()
        time.sleep(1)
        self.driver.find_element(*self.listar).click()



class Anexar_Guia(PageElement):
    anexar = (By.XPATH,'//*[@id="fupDoc"]')
    adicionar = (By.XPATH,'//*[@id="btnAdicionar"]')

    def injetar_guia(self):
        count = 0
        faturas_df = pd.read_excel(planilha)
        for index, linha in faturas_df.iterrows():
            if (f"{linha['Guia Anexada']}") == "Sim" or (f"{linha['Guia Anexada']}") == "Não Anexado":
                print(f"{linha['Guia Anexada']}")
                count = count + 1
                continue
            else:
                print('Guia pronta para ser anexada')
            count = count + 1
            
            print(f"{linha['Nro Guia GEAP']}")
            self.driver.get("https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/DetalhamentoGuia.aspx?IdGsp=" + f"{linha['Nro Guia GEAP']}")
            print('Entrando na guia')
            time.sleep(2)
            self.driver.find_element(*self.anexar).send_keys(linha["Caminho"])
            time.sleep(1)
            self.driver.find_element(*self.adicionar).click()
            time.sleep(2)
            try:
                id = self.driver.find_element(By.XPATH,'//*[@id="grvLista"]/tbody/tr[1]/th[1]').text
            except:
                dados = ['Não Anexado']
                df = pd.DataFrame(dados)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Planilha1', startrow= count, startcol=3, header=False, index=False)
                writer.save()
                continue

            if id == "Id Arquivo":
                print('Guia Anexada')                                                                   
                dados = ['Sim']
                df = pd.DataFrame(dados)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Planilha1', startrow= count, startcol=3, header=False, index=False)
                writer.save()
            else:
                print('Erro ao anexar guia')
                dados = ['Cancelado']
                df = pd.DataFrame(dados)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Planilha1', startrow= count, startcol=3, header=False, index=False)
                writer.save()
        self.driver.quit()

#------------------------------------------------------------------------------------------------------------------------------------------------
def anexar_guias():
    try:
        global planilha
        planilha = filedialog.askopenfilename()
        
        global url
        url = "https://www2.geap.com.br/auth/prestador.asp"

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128',
                'https': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128'
            }
        }

        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
        except:
            driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro inesperado' )

    try:
        login_page = Login(driver, url)
        login_page.open()

        login_page.exe_login(
            login = "23003723",
            senha = "amhpdf0073",
            cpf = "66661692120"
        )

        time.sleep(4)

        caminho(driver, url).exe_caminho()

        time.sleep(2)

        Anexar_Guia(driver, url).injetar_guia()
        tkinter.messagebox.showinfo( 'Automação GEAP Faturamento' , 'Envio de guias na GEAP Concluído 😎✌' )
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal da GEAP caiu 😢' )
        driver.quit()