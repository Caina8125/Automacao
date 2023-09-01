from tkinter import filedialog
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from openpyxl import load_workbook
import pyautogui
from abc import ABC
import pandas as pd
import time
import os
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
import tkinter

class PageElement(ABC):
    def __init__(self, driver, url='') -> None:
        self.driver = driver
        self.url = url
    def open(self):
        self.driver.get(self.url)

class Login(PageElement):
    multiusuario = (By.XPATH, '//*[@id="login_user"]/div[1]/label')
    prestador = (By.XPATH, '//*[@id="login_code"]')
    cpf = (By.XPATH, '//*[@id="login_cpf"]')
    senha = (By.XPATH, '//*[@id="login_password"]')
    login = (By.XPATH, '//*[@id="btnLogin"]')
    

    def exe_login(self, prestador, cpf, senha):
        self.driver.find_element(*self.multiusuario).click()
        self.driver.find_element(*self.prestador).send_keys(prestador)
        self.driver.find_element(*self.cpf).send_keys(cpf)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.login).click()
    
class Caminho(PageElement):
    portal_tiss = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[1]/nav/ul/li[21]/a')
    acompanhar_xml = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[2]/article/div[6]/div[4]/div[4]/div[4]/div/div[2]/ul/li[2]/a')

    def exe_caminho(self):
        self.driver.find_element(*self.portal_tiss).click()
        self.driver.find_element(*self.acompanhar_xml).click()
        self.driver.switch_to.window(self.driver.window_handles[1])

class Conferencia(PageElement):
    numero_envio = (By.XPATH, '//*[@id="NroProtocolo"]')
    listar = (By.XPATH, '//*[@id="MenuOptionReport"]')
    table = (By.XPATH, '//*[@id="objTableDetalhe"]')

    def pesquisar_envio(self):
        df = pd.read_excel(planilha)
        count_remessa = 0

        for index, linha in df.iterrows():
            count_remessa = count_remessa + 1
            if f"{linha['Observações']}" == "Pesquisado" or f"{linha['Observações']}" == "Erro ao pesquisar número de envio":
                continue
            numero_processo = (f"{linha['Nº Fatura']}").replace(".0", "")
            numero_envio = (f"{linha['Protocolo']}").replace(".0", "")
            nome_arquivo = numero_processo + '_' + numero_envio
            self.driver.find_element(*self.numero_envio).send_keys(numero_envio)
            self.driver.find_element(*self.listar).click()

            self.driver.get("https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/DetalhamentoEntrega.aspx?NroProtocolo=" + numero_envio)
            try:
                relatorio = self.driver.find_element(By.XPATH, '//*[@id="objTableDetalhe"]/tbody/tr/td').text
                if "Não existem registros na base da dados para o critério escolhido." in relatorio :
                    dado = {'Observações': ["Erro ao pesquisar número de envio"]}
                    df_remessa = pd.DataFrame(dado)
                    book = load_workbook(planilha) #alterar para o endereço do arquivo
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_remessa.to_excel(writer, 'Carta Remessa', startrow= count_remessa, startcol=2, header=False, index=False)
                    writer.save()
                    self.driver.get('https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/AcompanhamentoEntrega.aspx')
                    continue    
            except:
                pass
            table = self.driver.find_element(By.XPATH, '//*[@id="objTableDetalhe"]')
            tabela_html = table.get_attribute('outerHTML')
            df_processo = pd.read_html(tabela_html, header=1)[0]
            df_processo.to_excel("\\" + f"\\10.0.0.239\\Faturamento\\RAMON FAT 239\\GEAP - ROBÔS FAT\\Conferência de Anexos\\Respostas2\\{nome_arquivo}.xlsx", sheet_name=(f"{linha['Protocolo']}"), index=False)
            df_processo = pd.read_excel("\\" + f"\\10.0.0.239\\Faturamento\\RAMON FAT 239\\GEAP - ROBÔS FAT\\Conferência de Anexos\\Respostas2\\{nome_arquivo}.xlsx", header=1)
            df_processo["Arquivo"] = ""
            df_processo["Verificação"] = ""
            df_processo["Nº Fatura"] = numero_processo
            df_processo["Protocolo"] = numero_envio
            df_processo = df_processo.loc[:, ["Nº Fatura", "Protocolo", "Id Guia", "Guia Prestador", "Arquivo", "Verificação"]]
            df_processo.to_excel("\\" + f"\\10.0.0.239\\Faturamento\\RAMON FAT 239\\GEAP - ROBÔS FAT\\Conferência de Anexos\\Respostas2\\{nome_arquivo}.xlsx", sheet_name=(f"{linha['Protocolo']}"), index=False)
            print(df_processo)
            planilha2 = "\\" + f"\\10.0.0.239\\Faturamento\\RAMON FAT 239\\GEAP - ROBÔS FAT\\Conferência de Anexos\\Respostas2\\{nome_arquivo}.xlsx"
            count = 0

            for index2, linha2 in df_processo.iterrows():
                count = count + 1
                id_guia = (f"{linha2['Id Guia']}").replace(".0", "")
                self.driver.get("https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/DetalhamentoGuia.aspx?IdGsp=" + id_guia)
                try:
                    arquivo = self.driver.find_element(By.XPATH, '/html/body/form/div[3]/table/tbody/tr[6]/td/div/table/tbody/tr[2]/td[2]')
                    arquivo = self.driver.find_element(By.XPATH, '/html/body/form/div[3]/table/tbody/tr[6]/td/div/table/tbody/tr[2]/td[2]').text
                    numero_fatura = f"{linha2['Guia Prestador']}".replace(".0", "")
                    if numero_fatura in arquivo:
                        contem = "Ok"
                    else:
                        contem = "VERIFICAR"
                    dado = {'Arquivo': [arquivo], 'Verificação': [contem]}
                    df_guia = pd.DataFrame(dado)
                    book = load_workbook(planilha2)
                    writer = pd.ExcelWriter(planilha2, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_guia.to_excel(writer, (f"{linha['Protocolo']}"), startrow= count, startcol=4, header=False, index=False)
                    writer.save()
                except:
                    contem = "NÃO ANEXADO"
                    arquivo = "Não há arquivo"
                    dado = {'Arquivo': [arquivo], 'Verificação': [contem]}
                    df_guia = pd.DataFrame(dado)
                    book = load_workbook(planilha2)
                    writer = pd.ExcelWriter(planilha2, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_guia.to_excel(writer, (f"{linha['Protocolo']}"), startrow= count, startcol=4, header=False, index=False)
                    writer.save()
                    continue

            dado = {'Observações': ["Pesquisado"]}
            df_remessa = pd.DataFrame(dado)
            book = load_workbook(planilha) #alterar para o endereço do arquivo
            writer = pd.ExcelWriter(planilha, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_remessa.to_excel(writer, 'Carta Remessa', startrow= count_remessa, startcol=2, header=False, index=False)
            writer.save()
            self.driver.get('https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/AcompanhamentoEntrega.aspx')

        df_tudo = []
        lista_arquivos = os.listdir(r"\\10.0.0.239\Faturamento\RAMON FAT 239\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2")
        for arquivo in lista_arquivos:
            df_tudo.append(pd.read_excel("\\" + f"\\10.0.0.239\\Faturamento\\RAMON FAT 239\\GEAP - ROBÔS FAT\\Conferência de Anexos\\Respostas2\\" + arquivo))
            print(df_tudo)
        
        df_tudo = pd.concat(df_tudo, axis=0)
        df_tudo.to_excel(r"\\10.0.0.239\Faturamento\RAMON FAT 239\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2\Verificadas.xlsx", index=False)
        self.driver.quit()
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def conferencia():

    global planilha
    planilha = filedialog.askopenfilename()
    print(planilha)

    url = "https://www2.geap.com.br/auth/prestador.asp"
    
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')

    options = {
    'proxy': {
            'http': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128',
            'https': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128'
        }
    }
    try:
        servico = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
    except:
        driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)

    try:
        login_page = Login(driver, url)
        login_page.open()

        login_page.exe_login(
            prestador = "23003723",
            senha = "amhpdf0073",
            cpf = "66661692120"
        )

        time.sleep(2)

        Caminho(driver, url).exe_caminho()

        time.sleep(2)

        Conferencia(driver, url).pesquisar_envio()
        tkinter.messagebox.showinfo( 'Automação GEAP Conferência' , 'Pesquisa na GEAP Concluída 😎✌' )
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal da GEAP caiu 😢' )
        driver.quit()