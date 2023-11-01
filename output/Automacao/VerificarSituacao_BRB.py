import pandas as pd
import time
from abc import ABC
from tkinter import filedialog
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from Filtro_Faturamento import *
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from seleniumwire import webdriver
import tkinter

class PageElement(ABC):
    def __init__(self, driver, url=''):
        self.driver = driver
        self.url = url
    def open(self):
        self.driver.get(self.url)

class Login(PageElement):
    prestador_pj = (By.XPATH, '//*[@id="tipoAcesso"]/option[9]')
    usuario = (By.XPATH, '//*[@id="login-entry"]')
    senha = (By.XPATH, '//*[@id="password-entry"]')
    entrar = (By.XPATH, '//*[@id="BtnEntrar"]')

    def logar(self, usuario, senha):
        self.driver.find_element(*self.prestador_pj).click()
        self.driver.find_element(*self.usuario).send_keys(usuario)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.entrar).click()
        time.sleep(10)

class Caminho(PageElement):
    localizar_procedimentos = (By.XPATH, '//*[@id="menuPrincipal"]/div/div[2]/a/span')
    Alerta = (By.XPATH, '/html/body/ul/li/div/div[2]/button[2]')

    def exe_caminho(self):
        try:
            element = WebDriverWait(self.driver, 60.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menuPrincipal"]/div/div[4]/a/span')))
        except TimeoutException:
            print("Tempo de espera excedido. O site pode estar com delay ou fora do ar.")
            self.driver.quit()
        except Exception as e:
            print("Ocorreu um erro inesperado:", str(e))
            self.driver.quit()  
        self.driver.find_element(*self.localizar_procedimentos).click()
        time.sleep(2)
        self.driver.find_element(*self.Alerta).click()
        time.sleep(2)
    
    
class injetar_dados(PageElement):
    guia_op = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input') 
    buscar = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/span/span')
    
    def inserir_dados(self):
        faturas_df = pd.read_excel(planilha)
        count = 0
        guia_loc = ""
        for index, linha in faturas_df.iterrows():
            guia = int((f"{linha['Nº Guia']}").replace(".0",""))

            if (f"{linha['Pesquisado no Portal']}") == "Sim":
                print('Já foi feita a pesquisa desta autorização.')
                count = count + 1
                print(count)
                print('___________________________________________________________________________')
                continue

            if guia_loc == guia:
                count = count + 1
                continue
            
            guia_loc = int((f"{linha['Nº Guia']}").replace(".0",""))

            pesquisa = False
            while pesquisa == False:
                try:
                    self.driver.find_element(*self.guia_op).clear() 
                    self.driver.find_element(*self.guia_op).send_keys(guia)
                    self.driver.find_element(*self.buscar).click()
                    pesquisa = True
                except:
                    pass
            
            count = count + 1

            guia_df = faturas_df.loc[(faturas_df["Nº Guia"] == guia)]
            count2 = 0

            for index, linha2 in guia_df.iterrows():
                if (f"{linha['Pesquisado no Portal']}") == "Sim":
                    print('Já foi feita a pesquisa desta autorização.')
                    count2 = count2 + 1
                    print(count)
                    print('___________________________________________________________________________')
                    continue
                try:
                    user = False
                    while user == False:
                        try:
                            usuario = self.driver.find_element(By.XPATH, '//*[@id="menu_78B1E34CFC8E414D8EB4F83B534E4FB4"]').click()
                            user = True
                        except:
                            pass
                    situacao = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH,'//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div[3]/span'))).text
                    print(f"{guia} está {situacao}")
                    carteira = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[1]/strong[2]').text.replace("-", "")
                    carteira_planilha = "0" + f"{linha2['Matríc. Convênio']}".replace('.0', '')

                    if carteira == carteira_planilha:
                        matricula = 'Válida'
                    
                    else:
                        matricula = f'Inválida. Correta: {carteira}'

                    try:
                        todos = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[5]/a')
                        hover = ActionChains(self.driver).move_to_element(todos)
                        hover.perform()
                        procedimentos = self.driver.find_element(By.XPATH, '//*[@id="tooltip535066"]').text.replace('-', '').replace('.', '')

                    except:
                        procedimentos = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div').text.replace('-', '').replace('.', '')

                    procedimentos_planilha = f"{linha2['Procedimento']}".replace('.', '').replace('-', '')
                    
                    if procedimentos_planilha[0] == '1' and len(procedimentos_planilha) == 9:
                        data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                        df = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        count2 = count2 + 1
                        continue   

                    if procedimentos_planilha[0] == "0" or procedimentos_planilha[0] == "6" or procedimentos_planilha[0] == "7" or procedimentos_planilha[0] == "8":
                        data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                        df = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        count2 = count2 + 1
                        continue 

                    while procedimentos_planilha[0] == "9":
                        if procedimentos_planilha[0] == "9" and procedimentos_planilha[1] == "8":
                            print("Procedimento")
                            break

                        else:
                            data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                            df = pd.DataFrame(data)
                            book = load_workbook(planilha)
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                            writer.save()
                            count2 = count2 + 1
                            break

                    if procedimentos_planilha[0] == "9" and procedimentos_planilha[1] != "8":
                        continue

                    if procedimentos_planilha in procedimentos:
                        dados_proc = 'Ok'

                    else:
                        dados_proc = 'Não consta nesta autorização'
                    
                    data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': [dados_proc], 'Pesquisado no Portal': ['Sim']}
                    df = pd.DataFrame(data)
                    book = load_workbook(planilha)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    count2 = count2 + 1
                    print('___________________________________________________________________________')
                except:
                    situacao = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[1]').text
                    data = {'Situação': [situacao], 'Validação Carteira': [''], 'Validação Proc.': [''], 'Pesquisado no Portal': ['Sim']}
                    print(f"{guia}: {situacao}")
                    df = pd.DataFrame(data)
                    book = load_workbook(planilha)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    self.driver.find_element(*self.guia_op).clear()
                    ('___________________________________________________________________________')
        self.driver.quit()
            



#-------------------------------------------------------------------------
def verificar_brb():
    try:
        global planilha
        planilha = filedialog.askopenfilename()

        try:
            processar_planilha()
            remove()
        except:
            pass

        global url
        url = 'https://portal.saudebrb.com.br/GuiasTISS/Logon'

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128',
                'https': 'http://lucas.paz:Gsw2022&@10.0.0.230:3128'
            }
        }

        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
        except:
            driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)      
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro inesperado' ) 

    try:
        login_page = Login(driver, url)
        login_page.open()

        login_page.logar(
            usuario = '00735860000173_2',
            senha = '00735860000173'
            )

        time.sleep(4)

        Caminho(driver,url).exe_caminho()

        injetar_dados(driver,url).inserir_dados()

        print("Todas as guias foram pesquisadas com sucesso.")
        tkinter.messagebox.showinfo( 'Automação BRB Faturamento' , 'Verificação concluída 😎✌' )
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal do BRB caiu 😢' )
        driver.quit()






