import pandas as pd
import time
from tkinter import filedialog
import tkinter.messagebox
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from seleniumwire import webdriver
from page_element import PageElement

class capturar_protocolo(PageElement):
    acessar_portal = (By.XPATH, '/html/body/div[3]/div[3]/div[1]/form/div[1]/div[1]/div/a')
    usuario = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[1]/div/div[1]/div/input')
    senha = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[2]/div/div[1]/div[1]/input')
    entrar = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[2]/button/div[2]/div/div')
    fechar = (By.XPATH, '/html/body/div[4]/div[2]/div/div[3]/button')
    portal_tiss = (By.XPATH, '/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div[1]/div/div')
    fechar_modal = By.XPATH, '/html/body/div[3]/div[2]/div/button/span[2]/span/i'
    relatorios = By.XPATH, '/html/body/div[1]/div/div[1]/aside/div[1]/div[1]/div[12]/div/div[1]/div[3]/div'
    arquivo_tiss = By.XPATH, '/html/body/div[1]/div/div[1]/aside/div[1]/div[1]/div[12]/div/div[2]/div/div[7]/div[3]'
    inserir_protocolo = (By.XPATH, '//*[@id="NroProtocolo"]')
    baixar = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[2]/article/form/div/a')
    elemento2 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[5]')
    elemento3 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[3]/td[5]')

    def exe_login(self, senha, cpf):
        time.sleep(4)
        try:
            self.driver.find_element(*self.fechar).click()
        except:
            pass
        self.driver.find_element(*self.acessar_portal).click()
        time.sleep(2)
        try:
            self.driver.implicitly_wait(4)
            self.driver.find_element(*self.usuario).send_keys(cpf)
            time.sleep(2)
            self.driver.find_element(*self.senha).click()
            time.sleep(2)
            self.driver.find_element(*self.senha).send_keys(senha)
            time.sleep(2)
            self.driver.find_element(*self.entrar).click()
            time.sleep(2)
            self.driver.find_element(*self.portal_tiss)
            
        except:
            self.driver.implicitly_wait(180)
            self.driver.find_element(*self.portal_tiss)

    def exe_caminho(self):
        try:
            self.driver.find_element(*self.fechar_modal).click()
            time.sleep(1)
        except:
            pass
        self.driver.find_element(*self.relatorios).click()
        time.sleep(1)
        self.driver.find_element(*self.arquivo_tiss).click()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(2)

    def exe_capturar(self):

        global count
        global protocolo_plan
        count = 0

        faturas_df = pd.read_excel(planilha)
        for _ in range(0, 10):
            try:
                for _, linha in faturas_df.iterrows():
                    count = count + 1

                    protocolo_plan =  f"{linha['Protocolo']}".replace(".0","")
                    fatura_plan =  f"{linha['Faturas']}".replace(".0","")
                    if ((f"{linha['Verificação']}" == "Fatura encontrada") or (protocolo_plan == "Total Geral")):
                        print(f"{count}){protocolo_plan} : Fatura encontrada => {fatura_plan}")
                        continue

                    print(f"{count}) Buscando a fatura do Protocolo => {protocolo_plan}")
                    
                    self.driver.find_element(*self.inserir_protocolo).send_keys(protocolo_plan)
                    time.sleep(0.3)
                    self.driver.find_element(*self.baixar).click()
                    time.sleep(0.3)

                    #Bloco de código que insere o número da fatura na planilha
                    fatura_site = self.driver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[3]').text
                    n_fatura = [fatura_site]
                    df = pd.DataFrame(n_fatura)
                    book = load_workbook(planilha)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, 'Faturas', startrow = count, startcol = 1, header=False, index=False)
                    writer.save()

                    capturar_protocolo(driver,url).confere()

                    self.driver.back()
                    time.sleep(3)
                    self.driver.find_element(*self.inserir_protocolo).clear()
                    time.sleep(2)
                break
            
            except:
                self.open()
                login_page.exe_login(cpf = '66661692120', senha = "Amhp2023")
                self.exe_caminho()
                self.exe_capturar()

    def confere(self):

        plan = planilha
        plan_atualizada = pd.read_excel(plan)
        dados = pd.DataFrame(plan_atualizada)
        dados = dados.iloc[count - 1]
        try:
            fatura_plan = dados["Faturas"].astype(str).replace(".0","")
        except:
            fatura_plan = str(dados["Faturas"])
        try:
            protocolo_plan = dados["Protocolo"].astype(str).replace(".0","")
        except:
            protocolo_plan = str(dados["Protocolo"])

        fatura_site = self.driver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[3]').text
        protocolo_site  = self.driver.find_element(By.XPATH,'//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[1]').text

        if ((protocolo_plan == protocolo_site) & (fatura_site == fatura_plan)):
            confere = ["Fatura encontrada"]
            df = pd.DataFrame(confere)
            book = load_workbook(plan)
            writer = pd.ExcelWriter(plan, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'Faturas', startrow=count, startcol=5, header=False, index=False)
            writer.save()
        else:
            erro = ["Verificar"]
            df = pd.DataFrame(erro)
            book = load_workbook(plan)
            writer = pd.ExcelWriter(plan, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, 'Faturas', startrow= count, startcol=5, header=False, index=False)
            writer.save()
    
    

#-------------------------------------------------------------------------------------------------------------------------   

def iniciar(user, password):
    try:
        global planilha
        global url

        url = 'https://www2.geap.com.br/auth/prestador.asp'
        planilha = filedialog.askopenfilename()

        options = {
            'proxy' : {
                'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
            }
        }

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        global driver, login_page
        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options= options, options = chrome_options)
        except:
            driver = webdriver.Chrome(seleniumwire_options= options, options = chrome_options)

        login_page = capturar_protocolo(driver , url)

        login_page.open()

        login_page.exe_login(
            cpf = '66661692120',
            senha = "Amhp2023"
        )
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro inesperado' )
        
    try:
        login_page.exe_caminho()

        login_page.exe_capturar()
        tkinter.messagebox.showinfo( 'Automação GEAP Financeiro' , 'Busca de Faturas na GEAP Concluído 😎✌' )
    except Exception as e:
        tkinter.messagebox.showerror( 'Erro Automação' , f'Ocorreu uma exceção não tratada \n {e.__class__.__name__} - {e}' )